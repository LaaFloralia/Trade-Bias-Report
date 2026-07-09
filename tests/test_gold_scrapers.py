"""金ETFフロー (gold_etf) / 中銀ゴールド購入 (gold_cb) の
パース・集計・scraped_data 出力を検証する。
master_prompt.md セクション1.5 (ファンダ大局バイアス) の XAUUSD ドライバー入力。"""
from __future__ import annotations

import io
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from scrapers.gold_etf import _parse_archive_bytes  # noqa: E402
from scrapers.gold_cb import (  # noqa: E402
    _aggregate_monthly,
    _parse_sdmx_series,
    OZT_TO_TONNE,
)
from main import format_scraped_data  # noqa: E402


# ---------- gold_etf ----------

def _make_gld_xlsx(rows: list[tuple]) -> bytes:
    """(date, tonnes) 行から SPDR archive 風の XLSX を生成する。"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws_disc = wb.active
    ws_disc.title = "Disclaimer"
    ws_disc.append(["SPDR GOLD SHARES HISTORICAL DATA"])
    ws = wb.create_sheet("US GLD Historical Archive")
    ws.append(["Date", "Closing Price", "Daily Share Volume", "Tonnes of Gold"])
    for date, tonnes in rows:
        ws.append([date, "100.0", "1000", tonnes])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_gld_parse_basic_and_holiday_skip():
    rows = [
        ("30-Jun-2026", 1010.0),
        ("01-Jul-2026", 1005.36),
        ("02-Jul-2026", 1001.37),
        ("03-Jul-2026", "US Holiday"),  # 休場行はスキップされる
        ("06-Jul-2026", 1002.79),
        ("07-Jul-2026", 1002.51),
        ("08-Jul-2026", 1002.51),
    ]
    r = _parse_archive_bytes(_make_gld_xlsx(rows))
    assert r["error"] is None
    assert r["tonnes"] == 1002.51
    assert r["as_of_date"] == "2026-07-08"
    # 休場行を除いた 6 データ点 → 日次変化は新しい順に 5 件
    assert len(r["daily_flows"]) == 5
    assert r["daily_flows"][0] == {"date": "2026-07-08", "tonnes": 1002.51, "change_t": 0.0}
    assert r["daily_flows"][1]["change_t"] == -0.28
    # 5 営業日累計 = 最新 - 5 データ点前 (30-Jun の 1010.0)
    assert r["change_5d_t"] == round(1002.51 - 1010.0, 2)
    # 20 営業日分の履歴はない → None
    assert r["change_20d_t"] is None


def test_gld_streak_outflow():
    rows = [
        ("01-Jul-2026", 1010.0),
        ("02-Jul-2026", 1009.0),
        ("03-Jul-2026", 1008.0),
        ("06-Jul-2026", 1007.0),
    ]
    r = _parse_archive_bytes(_make_gld_xlsx(rows))
    assert r["streak_direction"] == "outflow"
    assert r["streak_days"] == 3


def test_gld_parse_error_when_no_header():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.append(["nothing", "here"])
    buf = io.BytesIO()
    wb.save(buf)
    r = _parse_archive_bytes(buf.getvalue())
    assert r["error"] is not None


# ---------- gold_cb ----------

def test_sdmx_parse_drops_zero_values():
    xml = (
        '<x><Series COUNTRY="KAZ" INDICATOR="IRFCLDT1_IRFCL56V_FTO">'
        '<Obs TIME_PERIOD="2026-M02" OBS_VALUE="11174283.905"/>'
        '<Obs TIME_PERIOD="2026-M03" OBS_VALUE="0.0"/>'
        '<Obs TIME_PERIOD="2026-M04" OBS_VALUE="11390654.938"/>'
        "</Series></x>"
    )
    per_country = _parse_sdmx_series(xml)
    assert "2026-M03" not in per_country["KAZ"]  # 0 は欠測扱い
    assert len(per_country["KAZ"]) == 2


def _ozt(tonnes: float) -> float:
    return tonnes / OZT_TO_TONNE


def test_aggregate_filters_artifacts():
    per_country = {
        # 正常国: 毎月 +10 t
        "POL": {"2026-M03": _ozt(600), "2026-M04": _ozt(610), "2026-M05": _ozt(620)},
        "CHN": {"2026-M03": _ozt(2300), "2026-M04": _ozt(2305), "2026-M05": _ozt(2310)},
        # レベル異常 (単位アーティファクト、米国超え): 系列ごと除外
        "AGO": {"2026-M03": 592_900_000, "2026-M04": 650_900_000, "2026-M05": 592_900_000},
        # 月次変化 ±100 t 超 (報告訂正): 当該月のみ除外
        "HND": {"2026-M03": _ozt(0.7), "2026-M04": _ozt(0.7), "2026-M05": _ozt(700)},
    }
    # MIN_REPORTERS チェックを通すため正常国を水増し
    for i in range(10):
        per_country[f"X{i:02d}"] = {
            "2026-M03": _ozt(50), "2026-M04": _ozt(50), "2026-M05": _ozt(50)
        }
    months = _aggregate_monthly(per_country)
    assert [m["period"] for m in months] == ["2026-M04", "2026-M05"]
    m4 = months[0]
    assert m4["net_tonnes"] == 15.0  # POL +10, CHN +5
    assert all(c != "AGO" for c, _ in m4["top_movers"])  # レベル異常は不在
    m5 = months[1]
    assert m5["net_tonnes"] == 15.0  # HND +699.3 は除外
    assert any(c == "HND" for c, _ in m5["excluded"])
    assert m5["reporters"] == 12  # HND を除く 12 カ国
    assert m5["partial"] is True  # 30 カ国未満は速報扱い


# ---------- format_scraped_data 出力 ----------

def test_format_emits_gold_sections():
    data = {
        "timestamp": "2026-07-09T09:00:00",
        "gold_etf": {
            "source": "SPDR Gold Shares (official API)", "symbol": "GLD",
            "tonnes": 1002.51, "as_of_date": "2026-07-08",
            "daily_flows": [
                {"date": "2026-07-08", "tonnes": 1002.51, "change_t": 0.0},
                {"date": "2026-07-07", "tonnes": 1002.51, "change_t": -0.28},
            ],
            "change_5d_t": -2.57, "change_20d_t": -17.41,
            "streak_days": 2, "streak_direction": "outflow",
            "error": None,
        },
        "gold_cb": {
            "source": "IMF IRFCL (monthly)", "symbol": "XAU_CB",
            "months": [
                {"period": "2026-M06", "net_tonnes": 1.6, "reporters": 11,
                 "partial": True, "top_movers": [], "excluded": []},
                {"period": "2026-M05", "net_tonnes": 32.4, "reporters": 56,
                 "partial": False, "top_movers": [("POL", 18.2), ("CHN", 10.0)],
                 "excluded": []},
            ],
            "cumulative_3m_t": 91.3, "regime": "net_buying",
            "as_of_date": "2026-M06",
            "note": "IRFCL 報告国ベースの集計。WGC 統計とは範囲・定義が異なる (レジーム判定用)",
            "error": None,
        },
    }
    text = format_scraped_data(data)
    assert "### 金ETFフロー (GLD 保有量)" in text
    assert "保有量: 1002.51 t (as_of 2026-07-08)" in text
    assert "5営業日累計: -2.57 t / 20営業日累計: -17.41 t" in text
    assert "2営業日連続流出" in text
    assert "### 中銀ゴールド購入 (IMF IRFCL 報告国ベース)" in text
    assert "[速報・報告国少]" in text
    assert "POL(ポーランド) +18.2" in text
    assert "レジーム: net_buying (中銀は買い越し基調)" in text


def test_format_emits_gold_errors():
    data = {
        "timestamp": "2026-07-09T09:00:00",
        "gold_etf": {"source": "SPDR Gold Shares (official API)", "error": "HTTP 500"},
        "gold_cb": {"source": "IMF IRFCL (monthly)", "error": "timeout"},
    }
    text = format_scraped_data(data)
    assert "### 金ETFフロー (GLD 保有量)" in text
    assert "取得不可（HTTP 500）" in text
    assert "取得不可（timeout）" in text
