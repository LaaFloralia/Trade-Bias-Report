"""金 ETF フロー スクレイパー (GLD 保有量)

ソース: SPDR Gold Shares 公式 historical-archive API (XLSX 全履歴)
  https://api.spdrgoldshares.com/api/v1/historical-archive?product=gld&exchange=NYSE&lang=en

BTC ETF フロー (btc_etf.py) の金版。GLD (世界最大の金現物 ETF) の保有トン数の
日次変化を機関投資家フローの代理変数として取得する。
master_prompt セクション1.5 (ファンダ大局バイアス) の XAUUSD ドライバー入力。

出力の解釈: 保有トン数の増加 = 機関の買い圧力 (create) / 減少 = 売り圧力 (redeem)。
"""

from __future__ import annotations

import asyncio
import io
import sys
from datetime import datetime
from pathlib import Path
from typing import Optional

sys.path.insert(0, str(Path(__file__).parent.parent))

import requests
from config import USER_AGENT

ARCHIVE_URL = (
    "https://api.spdrgoldshares.com/api/v1/historical-archive"
    "?product=gld&exchange=NYSE&lang=en"
)
HTTP_TIMEOUT = 60

# 何営業日ぶんの日次変化をレポートに載せるか / 累計ウィンドウ
DAILY_WINDOW = 5
TREND_WINDOW = 20


def _parse_date(raw: str) -> Optional[str]:
    """'08-Jul-2026' → '2026-07-08'。パース不能は None。"""
    try:
        return datetime.strptime(str(raw).strip(), "%d-%b-%Y").strftime("%Y-%m-%d")
    except ValueError:
        return None


def _parse_archive_bytes(data: bytes) -> dict:
    """historical-archive XLSX のバイト列から保有トン数の時系列を抽出・集計する。

    シート構成: ['Disclaimer', 'US GLD Historical Archive']
    データシートのヘッダー行: Date / Closing Price / ... / Tonnes of Gold
    休場日は全列 'US Holiday' 文字列になるためスキップする。
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True)
    sheet_name = next(
        (n for n in wb.sheetnames if "historical archive" in n.lower()),
        wb.sheetnames[-1],
    )
    ws = wb[sheet_name]

    date_idx: Optional[int] = None
    tonnes_idx: Optional[int] = None
    series: list[tuple[str, float]] = []  # (ISO date, tonnes) 古い順

    for row in ws.iter_rows(values_only=True):
        if date_idx is None:
            # ヘッダー行の探索
            cells = [str(c).strip().lower() if c is not None else "" for c in row]
            if "date" in cells and any("tonnes" in c for c in cells):
                date_idx = cells.index("date")
                tonnes_idx = next(i for i, c in enumerate(cells) if "tonnes" in c)
            continue
        if len(row) <= max(date_idx, tonnes_idx):
            continue
        iso = _parse_date(row[date_idx]) if row[date_idx] is not None else None
        if iso is None:
            continue
        try:
            tonnes = float(row[tonnes_idx])
        except (TypeError, ValueError):
            continue  # 'US Holiday' 等の非数値行
        series.append((iso, tonnes))

    if date_idx is None or not series:
        return {"error": "XLSX 内に Date/Tonnes ヘッダーまたはデータ行が見つからない"}

    # 日次変化 (直近 DAILY_WINDOW 営業日、新しい順)
    daily_flows = []
    for i in range(len(series) - 1, max(len(series) - 1 - DAILY_WINDOW, 0), -1):
        date, tonnes = series[i]
        change = tonnes - series[i - 1][1]
        daily_flows.append({"date": date, "tonnes": round(tonnes, 2), "change_t": round(change, 2)})

    def _window_change(n: int) -> Optional[float]:
        if len(series) <= n:
            return None
        return round(series[-1][1] - series[-1 - n][1], 2)

    # 連続方向 (0 変化・符号反転で打ち切り)
    streak_days = 0
    streak_direction = None
    for f in daily_flows:
        c = f["change_t"]
        if c == 0:
            break
        d = "inflow" if c > 0 else "outflow"
        if streak_direction is None:
            streak_direction = d
        if d != streak_direction:
            break
        streak_days += 1
    if streak_days == 0:
        streak_direction = None

    return {
        "tonnes": round(series[-1][1], 2),
        "as_of_date": series[-1][0],
        "daily_flows": daily_flows,
        "change_5d_t": _window_change(DAILY_WINDOW),
        "change_20d_t": _window_change(TREND_WINDOW),
        "streak_days": streak_days,
        "streak_direction": streak_direction,
        "error": None,
    }


def _fetch_archive() -> bytes:
    resp = requests.get(
        ARCHIVE_URL,
        headers={"User-Agent": USER_AGENT, "Accept": "*/*"},
        timeout=HTTP_TIMEOUT,
    )
    resp.raise_for_status()
    if resp.content[:2] != b"PK":
        raise ValueError(f"XLSX 以外のレスポンス (content-type={resp.headers.get('content-type')})")
    return resp.content


async def scrape_gold_etf() -> dict:
    """GLD 保有トン数の時系列を取得・集計する。

    Returns (metadata schema 準拠):
        {
            "source": "SPDR Gold Shares (official API)",
            "symbol": "GLD",
            "tonnes": float,             # 最新保有量 (トン)
            "as_of_date": "YYYY-MM-DD",
            "daily_flows": [{"date", "tonnes", "change_t"}],  # 新しい順 5 営業日
            "change_5d_t": float | None,   # 5 営業日累計変化
            "change_20d_t": float | None,  # 20 営業日累計変化 (中期トレンド)
            "streak_days": int,            # 同方向連続日数
            "streak_direction": "inflow" | "outflow" | None,
            "error": str | None,
        }
    """
    base = {"source": "SPDR Gold Shares (official API)", "symbol": "GLD"}
    print("  金ETF (GLD): SPDR historical-archive を取得中...")
    try:
        data = await asyncio.to_thread(_fetch_archive)
        parsed = await asyncio.to_thread(_parse_archive_bytes, data)
    except Exception as e:
        print(f"  [WARN] 金ETF (GLD): {e}")
        return {**base, "error": f"{type(e).__name__}: {e}"}
    if parsed.get("error"):
        print(f"  [WARN] 金ETF (GLD): {parsed['error']}")
        return {**base, **parsed}
    print(
        f"  [OK]    金ETF (GLD): {parsed['tonnes']} t (as_of {parsed['as_of_date']}, "
        f"5d {parsed['change_5d_t']:+} t)"
    )
    return {**base, **parsed}


if __name__ == "__main__":
    d = asyncio.run(scrape_gold_etf())
    if d.get("error"):
        print(f"Error: {d['error']}")
    else:
        print(f"GLD {d['tonnes']} t (as_of {d['as_of_date']})")
        for f in d["daily_flows"]:
            print(f"  {f['date']}: {f['tonnes']} t ({f['change_t']:+} t)")
        print(f"5d: {d['change_5d_t']:+} t / 20d: {d['change_20d_t']:+} t / streak: {d['streak_days']}日 {d['streak_direction']}")
